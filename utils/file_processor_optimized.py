"""
Procesador Optimizado de Archivos Grandes
==========================================

Sistema de procesamiento por streaming para archivos de 500GB+
- Streaming reading (no carga todo en memoria)
- Chunks de 50K-100K registros
- Garbage collection manual
- Progress tracking
- Memory efficient

Autor: Sistema SQL App
Fecha: 18 de octubre de 2025
"""

import pandas as pd
import gc
import logging
from typing import Generator, Dict, Any, Optional
from pathlib import Path
import json

logger = logging.getLogger(__name__)


class OptimizedFileProcessor:
    """
    Procesador optimizado para archivos grandes.
    
    Características:
    - Streaming reading (no carga todo en RAM)
    - Procesamiento por chunks
    - Memory efficient
    - Progress tracking
    """
    
    # Configuración de chunks por tipo de archivo
    CHUNK_SIZES = {
        'csv': 50000,      # 50K registros por chunk
        'excel': 10000,    # Excel es más pesado
        'txt': 100000,     # TXT es más ligero
        'json': 25000,     # JSON medio peso
        'parquet': 100000  # Parquet muy eficiente
    }
    
    def __init__(self, file_path: str, file_type: str):
        """
        Inicializa el procesador.
        
        Args:
            file_path: Ruta del archivo
            file_type: Tipo (csv, excel, txt, json, parquet)
        """
        self.file_path = Path(file_path)
        self.file_type = file_type.lower()
        self.chunk_size = self.CHUNK_SIZES.get(self.file_type, 50000)
        
        if not self.file_path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
        
        # Obtener tamaño del archivo
        self.file_size_bytes = self.file_path.stat().st_size
        self.file_size_gb = self.file_size_bytes / (1024**3)
        
        logger.info(f"📁 Archivo: {self.file_path.name} ({self.file_size_gb:.2f} GB)")
    
    def process_csv_streaming(self, encoding: str = 'utf-8', 
                             delimiter: str = ',',
                             on_progress: Optional[callable] = None) -> Generator[pd.DataFrame, None, None]:
        """
        Procesa CSV en modo streaming (archivos de 500GB+).
        
        Args:
            encoding: Codificación del archivo
            delimiter: Separador (,  ;  |  etc)
            on_progress: Callback para reportar progreso
        
        Yields:
            DataFrames con chunks de datos
        """
        logger.info(f"📥 Iniciando lectura streaming CSV: {self.file_path.name}")
        logger.info(f"🔧 Chunk size: {self.chunk_size:,} registros")
        
        try:
            chunk_number = 0
            total_rows = 0
            
            # pandas.read_csv con iterator=True NO carga todo en memoria
            with pd.read_csv(
                self.file_path,
                encoding=encoding,
                delimiter=delimiter,
                chunksize=self.chunk_size,
                low_memory=True,           # Optimización de memoria
                engine='c',                # Motor C (más rápido)
                iterator=True,
                on_bad_lines='skip'        # Skip líneas malformadas
            ) as reader:
                
                for chunk in reader:
                    chunk_number += 1
                    rows_in_chunk = len(chunk)
                    total_rows += rows_in_chunk
                    
                    logger.info(f"📦 Chunk #{chunk_number}: {rows_in_chunk:,} registros (Total: {total_rows:,})")
                    
                    # Callback de progreso
                    if on_progress:
                        on_progress(chunk_number, rows_in_chunk, total_rows)
                    
                    # Yield del chunk
                    yield chunk
                    
                    # Limpiar memoria después de cada chunk
                    del chunk
                    gc.collect()
            
            logger.info(f"✅ CSV procesado: {total_rows:,} registros en {chunk_number} chunks")
            
        except Exception as e:
            logger.error(f"❌ Error procesando CSV: {str(e)}")
            raise
    
    def process_excel_streaming(self, sheet_name: str = 0,
                               on_progress: Optional[callable] = None) -> Generator[pd.DataFrame, None, None]:
        """
        Procesa Excel en modo streaming (archivos grandes).
        
        IMPORTANTE: Excel no tiene streaming nativo tan eficiente como CSV.
        Para archivos 500GB+ se recomienda convertir a CSV primero.
        
        Args:
            sheet_name: Nombre o índice de la hoja
            on_progress: Callback de progreso
        
        Yields:
            DataFrames con chunks de datos
        """
        logger.info(f"📊 Procesando Excel: {self.file_path.name}")
        
        if self.file_size_gb > 5:
            logger.warning(f"⚠️ Archivo Excel muy grande ({self.file_size_gb:.2f} GB)")
            logger.warning(f"💡 Recomendación: Convertir a CSV para mejor performance")
        
        try:
            # Opción 1: Usar openpyxl en modo read_only (más eficiente)
            from openpyxl import load_workbook
            
            wb = load_workbook(
                filename=str(self.file_path),
                read_only=True,    # Modo streaming
                data_only=True     # Solo valores, no fórmulas
            )
            
            ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]
            
            # Leer por chunks manualmente
            chunk_data = []
            chunk_number = 0
            total_rows = 0
            
            # Obtener headers
            headers = [cell.value for cell in next(ws.rows)]
            
            for row in ws.rows:
                row_data = [cell.value for cell in row]
                chunk_data.append(row_data)
                
                if len(chunk_data) >= self.chunk_size:
                    # Crear DataFrame del chunk
                    df_chunk = pd.DataFrame(chunk_data, columns=headers)
                    chunk_number += 1
                    total_rows += len(df_chunk)
                    
                    logger.info(f"📦 Chunk #{chunk_number}: {len(df_chunk):,} registros")
                    
                    if on_progress:
                        on_progress(chunk_number, len(df_chunk), total_rows)
                    
                    yield df_chunk
                    
                    # Limpiar
                    chunk_data = []
                    del df_chunk
                    gc.collect()
            
            # Último chunk (residual)
            if chunk_data:
                df_chunk = pd.DataFrame(chunk_data, columns=headers)
                chunk_number += 1
                total_rows += len(df_chunk)
                
                logger.info(f"📦 Chunk final #{chunk_number}: {len(df_chunk):,} registros")
                
                if on_progress:
                    on_progress(chunk_number, len(df_chunk), total_rows)
                
                yield df_chunk
            
            wb.close()
            logger.info(f"✅ Excel procesado: {total_rows:,} registros")
            
        except Exception as e:
            logger.error(f"❌ Error procesando Excel: {str(e)}")
            raise
    
    def process_txt_streaming(self, encoding: str = 'utf-8',
                             delimiter: str = None,
                             on_progress: Optional[callable] = None) -> Generator[pd.DataFrame, None, None]:
        """
        Procesa archivos TXT delimitados (similar a CSV).
        
        Args:
            encoding: Codificación
            delimiter: Separador (None = whitespace)
            on_progress: Callback
        
        Yields:
            DataFrames con chunks
        """
        logger.info(f"📄 Procesando TXT: {self.file_path.name}")
        
        # TXT es similar a CSV
        yield from self.process_csv_streaming(
            encoding=encoding,
            delimiter=delimiter or r'\s+',  # Whitespace por defecto
            on_progress=on_progress
        )
    
    def process_json_streaming(self, on_progress: Optional[callable] = None) -> Generator[pd.DataFrame, None, None]:
        """
        Procesa JSON en modo streaming (para archivos grandes).
        
        Soporta:
        - JSON Lines (.jsonl) - Recomendado para archivos grandes
        - JSON Array - Menos eficiente
        
        Args:
            on_progress: Callback
        
        Yields:
            DataFrames con chunks
        """
        logger.info(f"📋 Procesando JSON: {self.file_path.name}")
        
        try:
            # Detectar formato
            with open(self.file_path, 'r', encoding='utf-8') as f:
                first_char = f.read(1)
            
            if first_char == '[':
                # JSON Array - usar ijson para streaming
                logger.info("📋 Formato: JSON Array (usando streaming)")
                yield from self._process_json_array_streaming(on_progress)
            else:
                # JSON Lines - lectura línea por línea
                logger.info("📋 Formato: JSON Lines (óptimo)")
                yield from self._process_jsonl_streaming(on_progress)
                
        except Exception as e:
            logger.error(f"❌ Error procesando JSON: {str(e)}")
            raise
    
    def _process_jsonl_streaming(self, on_progress: Optional[callable] = None) -> Generator[pd.DataFrame, None, None]:
        """Procesa JSON Lines (una línea = un objeto JSON)."""
        chunk_data = []
        chunk_number = 0
        total_rows = 0
        
        with open(self.file_path, 'r', encoding='utf-8') as f:
            for line_number, line in enumerate(f, 1):
                try:
                    obj = json.loads(line.strip())
                    chunk_data.append(obj)
                    
                    if len(chunk_data) >= self.chunk_size:
                        df_chunk = pd.DataFrame(chunk_data)
                        chunk_number += 1
                        total_rows += len(df_chunk)
                        
                        logger.info(f"📦 Chunk #{chunk_number}: {len(df_chunk):,} registros")
                        
                        if on_progress:
                            on_progress(chunk_number, len(df_chunk), total_rows)
                        
                        yield df_chunk
                        
                        chunk_data = []
                        del df_chunk
                        gc.collect()
                        
                except json.JSONDecodeError:
                    logger.warning(f"⚠️ Línea {line_number} inválida, skip")
                    continue
        
        # Último chunk
        if chunk_data:
            df_chunk = pd.DataFrame(chunk_data)
            chunk_number += 1
            total_rows += len(df_chunk)
            
            if on_progress:
                on_progress(chunk_number, len(df_chunk), total_rows)
            
            yield df_chunk
        
        logger.info(f"✅ JSON Lines procesado: {total_rows:,} registros")
    
    def _process_json_array_streaming(self, on_progress: Optional[callable] = None) -> Generator[pd.DataFrame, None, None]:
        """Procesa JSON Array usando ijson (streaming parser)."""
        try:
            import ijson
        except ImportError:
            logger.error("❌ ijson no instalado. Ejecuta: pip install ijson")
            raise ImportError("ijson requerido para streaming de JSON Arrays grandes")
        
        chunk_data = []
        chunk_number = 0
        total_rows = 0
        
        with open(self.file_path, 'rb') as f:
            # ijson.items() lee item por item sin cargar todo
            parser = ijson.items(f, 'item')
            
            for obj in parser:
                chunk_data.append(obj)
                
                if len(chunk_data) >= self.chunk_size:
                    df_chunk = pd.DataFrame(chunk_data)
                    chunk_number += 1
                    total_rows += len(df_chunk)
                    
                    logger.info(f"📦 Chunk #{chunk_number}: {len(df_chunk):,} registros")
                    
                    if on_progress:
                        on_progress(chunk_number, len(df_chunk), total_rows)
                    
                    yield df_chunk
                    
                    chunk_data = []
                    del df_chunk
                    gc.collect()
        
        # Último chunk
        if chunk_data:
            df_chunk = pd.DataFrame(chunk_data)
            chunk_number += 1
            total_rows += len(df_chunk)
            
            if on_progress:
                on_progress(chunk_number, len(df_chunk), total_rows)
            
            yield df_chunk
        
        logger.info(f"✅ JSON Array procesado: {total_rows:,} registros")
    
    def process_parquet_streaming(self, on_progress: Optional[callable] = None) -> Generator[pd.DataFrame, None, None]:
        """
        Procesa Parquet en modo streaming (ULTRA EFICIENTE).
        
        Parquet es el formato MÁS EFICIENTE para archivos grandes:
        - Compresión nativa
        - Columnar storage
        - Streaming nativo
        - 10x más rápido que CSV
        
        Args:
            on_progress: Callback
        
        Yields:
            DataFrames con chunks
        """
        logger.info(f"🚀 Procesando Parquet: {self.file_path.name}")
        logger.info(f"💡 Parquet es el formato MÁS EFICIENTE para archivos grandes")
        
        try:
            import pyarrow.parquet as pq
        except ImportError:
            logger.error("❌ pyarrow no instalado. Ejecuta: pip install pyarrow")
            raise ImportError("pyarrow requerido para Parquet")
        
        try:
            # Abrir archivo Parquet
            parquet_file = pq.ParquetFile(str(self.file_path))
            
            chunk_number = 0
            total_rows = 0
            
            # Leer por batches (streaming nativo de Parquet)
            for batch in parquet_file.iter_batches(batch_size=self.chunk_size):
                df_chunk = batch.to_pandas()
                chunk_number += 1
                total_rows += len(df_chunk)
                
                logger.info(f"📦 Chunk #{chunk_number}: {len(df_chunk):,} registros")
                
                if on_progress:
                    on_progress(chunk_number, len(df_chunk), total_rows)
                
                yield df_chunk
                
                del df_chunk, batch
                gc.collect()
            
            logger.info(f"✅ Parquet procesado: {total_rows:,} registros")
            
        except Exception as e:
            logger.error(f"❌ Error procesando Parquet: {str(e)}")
            raise
    
    def get_estimated_rows(self) -> Optional[int]:
        """
        Estima la cantidad de filas sin leer todo el archivo.
        
        Returns:
            Estimación de filas o None si no se puede estimar
        """
        try:
            if self.file_type == 'csv':
                # Contar líneas rápidamente
                with open(self.file_path, 'rb') as f:
                    line_count = sum(1 for _ in f)
                return line_count - 1  # -1 por header
            
            elif self.file_type == 'parquet':
                import pyarrow.parquet as pq
                parquet_file = pq.ParquetFile(str(self.file_path))
                return parquet_file.metadata.num_rows
            
            else:
                # Para otros formatos, no estimamos (requiere lectura)
                return None
                
        except Exception as e:
            logger.warning(f"⚠️ No se pudo estimar filas: {str(e)}")
            return None


def detect_file_encoding(file_path: str, sample_size: int = 10000) -> str:
    """
    Detecta la codificación de un archivo de texto.
    
    Args:
        file_path: Ruta del archivo
        sample_size: Bytes a leer para detección
    
    Returns:
        Codificación detectada (utf-8, latin1, etc.)
    """
    try:
        import chardet
    except ImportError:
        logger.warning("⚠️ chardet no instalado, asumiendo utf-8")
        return 'utf-8'
    
    with open(file_path, 'rb') as f:
        sample = f.read(sample_size)
    
    result = chardet.detect(sample)
    encoding = result['encoding']
    confidence = result['confidence']
    
    logger.info(f"🔍 Encoding detectado: {encoding} (confianza: {confidence:.0%})")
    
    return encoding or 'utf-8'


def get_optimal_chunk_size(file_size_gb: float, available_ram_gb: float = 4) -> int:
    """
    Calcula el chunk size óptimo según tamaño de archivo y RAM disponible.
    
    Args:
        file_size_gb: Tamaño del archivo en GB
        available_ram_gb: RAM disponible en GB
    
    Returns:
        Chunk size óptimo en registros
    """
    # Reglas heurísticas
    if file_size_gb < 1:
        return 100000  # 100K registros
    elif file_size_gb < 10:
        return 50000   # 50K registros
    elif file_size_gb < 100:
        return 25000   # 25K registros
    else:
        # Archivos 100GB+: chunks muy pequeños
        return 10000   # 10K registros
